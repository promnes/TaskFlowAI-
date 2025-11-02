#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
تنسيق ملفات Excel بشكل احترافي
Professional Excel Formatting System
"""

import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import os

class ExcelFormatter:
    def __init__(self):
        # ألوان النظام
        self.colors = {
            'primary': '1F4E79',      # أزرق داكن
            'secondary': '2E75B6',    # أزرق متوسط
            'success': '70AD47',      # أخضر
            'warning': 'FFC000',      # أصفر
            'danger': 'C5504B',       # أحمر
            'light': 'F2F2F2',       # رمادي فاتح
            'white': 'FFFFFF',        # أبيض
            'text': '2D2D2D'          # نص داكن
        }
        
        # خطوط النظام
        self.fonts = {
            'header': Font(name='Calibri', size=14, bold=True, color=self.colors['white']),
            'subheader': Font(name='Calibri', size=12, bold=True, color=self.colors['primary']),
            'body': Font(name='Calibri', size=10, color=self.colors['text']),
            'highlight': Font(name='Calibri', size=10, bold=True, color=self.colors['primary'])
        }
        
        # حدود الجدول
        thin_border = Side(border_style="thin", color=self.colors['primary'])
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def create_professional_workbook(self, filename='DUX_Professional_Report.xlsx'):
        """إنشاء مصنف Excel احترافي"""
        wb = openpyxl.Workbook()
        
        # حذف الورقة الافتراضية
        wb.remove(wb.active)
        
        # إضافة الأوراق المطلوبة
        sheets = [
            ('المستخدمين', 'users.csv'),
            ('المعاملات', 'transactions.csv'),
            ('الشكاوى', 'complaints.csv'),
            ('الشركات', 'companies.csv'),
            ('وسائل الدفع', 'payment_methods.csv'),
            ('الإحصائيات', 'summary')
        ]
        
        for sheet_name, csv_file in sheets:
            if csv_file == 'summary':
                self.create_summary_sheet(wb, sheet_name)
            else:
                self.create_formatted_sheet(wb, sheet_name, csv_file)
        
        # حفظ الملف
        wb.save(filename)
        return filename
    
    def create_formatted_sheet(self, wb, sheet_name, csv_file):
        """إنشاء ورقة مُنسقة من ملف CSV"""
        if not os.path.exists(csv_file):
            return
        
        ws = wb.create_sheet(title=sheet_name)
        
        try:
            with open(csv_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                data = list(reader)
            
            if not data:
                return
            
            # إضافة البيانات
            for row_idx, row in enumerate(data, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # تنسيق الرأس
                    if row_idx == 1:
                        self.format_header_cell(cell)
                    else:
                        self.format_data_cell(cell, csv_file, col_idx)
            
            # تطبيق التنسيق الاحترافي
            self.apply_professional_formatting(ws, len(data[0]) if data else 0, len(data))
            
            # إضافة جدول
            if len(data) > 1:
                self.add_table_style(ws, len(data[0]), len(data))
            
        except Exception as e:
            print(f"خطأ في تنسيق {csv_file}: {e}")
    
    def format_header_cell(self, cell):
        """تنسيق خلايا الرأس"""
        cell.font = self.fonts['header']
        cell.fill = PatternFill(start_color=self.colors['primary'], end_color=self.colors['primary'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.border
    
    def format_data_cell(self, cell, csv_file, col_idx):
        """تنسيق خلايا البيانات"""
        cell.font = self.fonts['body']
        cell.border = self.border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # تنسيق خاص حسب نوع البيانات
        if csv_file == 'transactions.csv':
            self.format_transaction_cell(cell, col_idx)
        elif csv_file == 'users.csv':
            self.format_user_cell(cell, col_idx)
        elif csv_file == 'complaints.csv':
            self.format_complaint_cell(cell, col_idx)
    
    def format_transaction_cell(self, cell, col_idx):
        """تنسيق خاص للمعاملات"""
        value = str(cell.value).lower() if cell.value else ''
        
        # تلوين حالة المعاملة (عمود status)
        if col_idx == 10:  # عمود الحالة
            if value == 'approved':
                cell.fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['white'])
            elif value == 'rejected':
                cell.fill = PatternFill(start_color=self.colors['danger'], end_color=self.colors['danger'], fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['white'])
            elif value == 'pending':
                cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['text'])
        
        # تنسيق نوع المعاملة (عمود type)
        elif col_idx == 5:
            if value == 'deposit':
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['success'])
            elif value == 'withdraw':
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['danger'])
    
    def format_user_cell(self, cell, col_idx):
        """تنسيق خاص للمستخدمين"""
        value = str(cell.value).lower() if cell.value else ''
        
        # تلوين حالة الحظر (عمود is_banned)
        if col_idx == 7:
            if value == 'yes':
                cell.fill = PatternFill(start_color=self.colors['danger'], end_color=self.colors['danger'], fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['white'])
            else:
                cell.fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['white'])
    
    def format_complaint_cell(self, cell, col_idx):
        """تنسيق خاص للشكاوى"""
        value = str(cell.value).lower() if cell.value else ''
        
        # تلوين حالة الشكوى (عمود status)
        if col_idx == 5:
            if value == 'resolved':
                cell.fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['white'])
            elif value == 'pending':
                cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['text'])
    
    def apply_professional_formatting(self, ws, cols, rows):
        """تطبيق التنسيق الاحترافي"""
        # تحديد عرض الأعمدة
        for col in range(1, cols + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # تجميد الصف الأول
        ws.freeze_panes = 'A2'
        
        # ضبط اتجاه النص للعربية
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # التحقق من وجود نص عربي
                    if any('\u0600' <= char <= '\u06FF' for char in cell.value):
                        cell.alignment = Alignment(horizontal='right', vertical='center', text_rotation=0)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def add_table_style(self, ws, cols, rows):
        """إضافة نمط جدول احترافي"""
        # تحديد نطاق الجدول
        table_range = f"A1:{get_column_letter(cols)}{rows}"
        
        # إنشاء الجدول
        table = Table(displayName=f"Table_{ws.title}", ref=table_range)
        
        # تطبيق نمط جدول احترافي
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        # إضافة الجدول للورقة
        ws.add_table(table)
    
    def create_summary_sheet(self, wb, sheet_name):
        """إنشاء ورقة الإحصائيات"""
        ws = wb.create_sheet(title=sheet_name)
        
        # بيانات الإحصائيات
        stats = self.calculate_statistics()
        
        # إضافة عنوان رئيسي
        ws['A1'] = 'تقرير إحصائيات نظام DUX المالي'
        ws['A1'].font = Font(name='Calibri', size=18, bold=True, color=self.colors['primary'])
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        
        # إضافة تاريخ التقرير
        ws['A2'] = f'تاريخ التقرير: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws['A2'].font = self.fonts['subheader']
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:E2')
        
        # إضافة الإحصائيات
        row = 4
        for category, data in stats.items():
            # عنوان القسم
            ws[f'A{row}'] = category
            ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color=self.colors['primary'])
            ws[f'A{row}'].fill = PatternFill(start_color=self.colors['light'], end_color=self.colors['light'], fill_type='solid')
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # بيانات القسم
            for key, value in data.items():
                ws[f'B{row}'] = key
                ws[f'C{row}'] = value
                ws[f'B{row}'].font = self.fonts['body']
                ws[f'C{row}'].font = self.fonts['highlight']
                row += 1
            
            row += 1  # مسافة بين الأقسام
        
        # تنسيق الأعمدة
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def calculate_statistics(self):
        """حساب الإحصائيات"""
        stats = {
            'إحصائيات المستخدمين': {},
            'إحصائيات المعاملات': {},
            'إحصائيات الشكاوى': {},
            'إحصائيات الشركات': {}
        }
        
        try:
            # إحصائيات المستخدمين
            if os.path.exists('users.csv'):
                with open('users.csv', 'r', encoding='utf-8-sig') as f:
                    users = list(csv.DictReader(f))
                    stats['إحصائيات المستخدمين'] = {
                        'إجمالي المستخدمين': len(users),
                        'المستخدمين النشطين': len([u for u in users if u.get('is_banned', 'no').lower() != 'yes']),
                        'المستخدمين المحظورين': len([u for u in users if u.get('is_banned', 'no').lower() == 'yes']),
                        'مستخدمي العربية': len([u for u in users if u.get('language', 'ar') == 'ar']),
                        'مستخدمي الإنجليزية': len([u for u in users if u.get('language', 'ar') == 'en'])
                    }
            
            # إحصائيات المعاملات
            if os.path.exists('transactions.csv'):
                with open('transactions.csv', 'r', encoding='utf-8-sig') as f:
                    transactions = list(csv.DictReader(f))
                    
                    approved = [t for t in transactions if t.get('status') == 'approved']
                    rejected = [t for t in transactions if t.get('status') == 'rejected']
                    pending = [t for t in transactions if t.get('status') == 'pending']
                    deposits = [t for t in transactions if t.get('type') == 'deposit']
                    withdrawals = [t for t in transactions if t.get('type') == 'withdraw']
                    
                    stats['إحصائيات المعاملات'] = {
                        'إجمالي المعاملات': len(transactions),
                        'المعاملات المُوافقة': len(approved),
                        'المعاملات المرفوضة': len(rejected),
                        'المعاملات المعلقة': len(pending),
                        'طلبات الإيداع': len(deposits),
                        'طلبات السحب': len(withdrawals),
                        'معدل الموافقة': f"{(len(approved)/len(transactions)*100):.1f}%" if transactions else "0%"
                    }
            
            # إحصائيات الشكاوى
            if os.path.exists('complaints.csv'):
                with open('complaints.csv', 'r', encoding='utf-8-sig') as f:
                    complaints = list(csv.DictReader(f))
                    
                    resolved = [c for c in complaints if c.get('status') == 'resolved']
                    pending = [c for c in complaints if c.get('status') == 'pending']
                    
                    stats['إحصائيات الشكاوى'] = {
                        'إجمالي الشكاوى': len(complaints),
                        'الشكاوى المحلولة': len(resolved),
                        'الشكاوى المعلقة': len(pending),
                        'معدل الحل': f"{(len(resolved)/len(complaints)*100):.1f}%" if complaints else "0%"
                    }
            
            # إحصائيات الشركات
            if os.path.exists('companies.csv'):
                with open('companies.csv', 'r', encoding='utf-8-sig') as f:
                    companies = list(csv.DictReader(f))
                    
                    active = [c for c in companies if c.get('is_active', '').lower() == 'active']
                    both_type = [c for c in companies if c.get('type') == 'both']
                    deposit_only = [c for c in companies if c.get('type') == 'deposit']
                    withdraw_only = [c for c in companies if c.get('type') == 'withdraw']
                    
                    stats['إحصائيات الشركات'] = {
                        'إجمالي الشركات': len(companies),
                        'الشركات النشطة': len(active),
                        'شركات الإيداع والسحب': len(both_type),
                        'شركات الإيداع فقط': len(deposit_only),
                        'شركات السحب فقط': len(withdraw_only)
                    }
        
        except Exception as e:
            print(f"خطأ في حساب الإحصائيات: {e}")
        
        return stats

# مثال للاستخدام
if __name__ == "__main__":
    formatter = ExcelFormatter()
    filename = formatter.create_professional_workbook()
    print(f"تم إنشاء الملف الاحترافي: {filename}")